import re
import os
import json
import math
import warnings
import numpy as np
from interval import Interval
from decimal import *
import pandas as pd
import chardet
import xlsxwriter
import openpyxl 
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Alignment



with open("config.json", "r") as jn:
    config = json.loads(jn.read())
    
folders = ['infile', 'outfile','lastfile','otherfile']

gap_file = "gaps_data\\"
gap_dataf = os.listdir(gap_file)

for folder in folders:
    if not os.path.isdir(folder):
        os.mkdir(folder)

infile="infile\\"
outfile="outfile\\"
lastfile="lastfile\\"
otherfile = "otherfile\\"
print("\n")
print("Processing》》》》》》")
print("\n")



param_region_1 = 0

result_df_count = pd.DataFrame()

pd.set_option('display.max_columns', 50)  
pd.set_option('display.max_rows', 100)   

#Create a list of chromosomes 1 to Y
chromosomes = [f'{i:02}' for i in range(1, 23)] + ['X', 'Y']


# Traverse Chromosome List
for chromosome in chromosomes:
    #print(chromosome)
    chr_parm = chromosome #chr
    #Obtain file encoding format
    def get_encoding(file):
        with open(file,'rb') as f:
            file_format = chardet.detect(f.read())['encoding']
            #print(file_format)
            return file_format

    #Convert CSV files to xlsx files and process the data format in the files
    def csv_to_xlsx(csv_file, xlsx_file):
        # Read CSV file and obtain column names
        df = pd.read_csv(csv_file,header=None,low_memory=False,encoding=get_encoding(csv_file))
        #print(df)
        df.columns = range(df.shape[1])
        data_head = df.loc[0]
        #print(df)
        column_names = df.columns.tolist()
        #print(column_names)
        # Select the column that needs to modify the format
        #cols_to_change = ["1","2","4","5","8","9","11","12","15","16","18","19","21","22","24","25"]
        # change formatting
        cols = [1,2,4,5,6,8,9,11,12,16,19]
        df_slice = df.iloc[1:, df.columns.isin(cols)]
        #df_slice = df_slice.applymap(lambda x: float(x))
        #print(df_slice)
        for col in df_slice.columns:
            df_slice[col] = pd.to_numeric(df_slice[col], errors='coerce', downcast='integer').fillna(df_slice[col]) #忽略NAN数据和非数字数据
        df[cols] = df_slice
        df.loc[0] = data_head
        # Convert to XLSX file and save
        writer = pd.ExcelWriter(xlsx_file, engine='xlsxwriter')
        df.to_excel(writer,index=False,header=False, sheet_name='Sheet1')
        writer.save()



    def df_insert(df, n, arr, ignore_index=True):
        new_row = pd.DataFrame([arr], columns=df.columns)
        df1 = df.iloc[:n]
        df2 = df.iloc[n:]
        df0 = pd.concat([df1, new_row, df2], ignore_index=ignore_index)
        return df0

    # Continuous processing of data
    def process_data(data):
        new_data = pd.DataFrame(columns=['index_no','[S1]', '[E1]', '[S2]', '[E2]', '[LEN1]','[LEN2]','[%IDY]','Extra_Column'])
        new_rows = []

        # Process data before the first row
        first_S1, first_S2 = data.loc[0, '[S1]'], data.loc[0, '[S2]']
        if first_S1 > 1 or first_S2 > 1:
            new_row = {
                '[S1]': 1,
                '[E1]': first_S1 - 1,
                '[S2]': 1,
                '[E2]': first_S2 - 1,
                '[LEN1]': None,
                '[LEN2]': None,
                '[%IDY]': None,
                'Extra_Column': 'Inserted'
            }
            new_rows.append(new_row)

        # Process Subsequent Lines
        for i in range(len(data)):
            S1_data, E1_data, S2_data, E2_data = data.loc[i, '[S1]'], data.loc[i, '[E1]'], data.loc[i, '[S2]'], data.loc[i, '[E2]']

            # Inserting intervals to make data coherent
            if i > 0:
                prev_E1, prev_E2 = data.loc[i-1, '[E1]'], data.loc[i-1, '[E2]']
                chm13_diff = S1_data - prev_E1 - 1
                grch38_diff = S2_data - prev_E2 - 1
                #if chm13_diff > 0 or grch38_diff > 0:
                chm13_interval_start = prev_E1 + 1
                chm13_interval_end = S1_data - 1
                grch38_interval_start = prev_E2 + 1
                grch38_interval_end = S2_data - 1
                new_row = {
                    '[S1]': chm13_interval_start,
                    '[E1]': chm13_interval_end,
                    '[S2]': grch38_interval_start,
                    '[E2]': grch38_interval_end,
                    '[LEN1]': None,
                    '[LEN2]': None,
                    '[%IDY]': None,
                    'Extra_Column': 'Inserted'
                }
                new_rows.append(new_row)

            # Insert current row data
            new_rows.append(data.iloc[i].to_dict())

        # Add new row data to a new DataFrame
        new_data = pd.concat([new_data, pd.DataFrame(new_rows)], ignore_index=True)

        return new_data



    def find_nearest_position(value, start, end):
        # Calculate the difference between the given value and the starting position
        diff_start = value - start
        # Calculate the difference between the given value and the end position
        diff_end = value - end

        if abs(diff_start) < abs(diff_end):
            return diff_start, 'start'  
        else:
            return diff_end, 'end' 

    mummer_data = pd.read_csv(infile+"c13_g38_chr"+chr_parm+"_99_1kf.csv",encoding=get_encoding(infile+"c13_g38_chr"+chr_parm+"_99_1kf.csv"))
    mummer_data['index_no'] = range(len(mummer_data))
    # print(mummer_data)
    deal_mummer_data = process_data(mummer_data)

        
    deal_mummer_data.to_csv(infile+"c13_g38_chr"+chr_parm+"_99_1kf_deal.csv", index=False) #

    CHM13_peak_data = pd.read_csv(infile+"HS-"+chr_parm+"_HP_count.csv",encoding=get_encoding(infile+"HS-"+chr_parm+"_HP_count.csv"))


    CHM13_peak_data['Peak_name'] = CHM13_peak_data['Peak_name'].str.replace('HS', 'CHM13')  
    CHM13_peak_data['Peak_name'] = CHM13_peak_data['Peak_name'].str.replace('HDAP', 'HMDP')  
    CHM13_peak_data['Peak_name'] = CHM13_peak_data['Peak_name'].str.replace('MDAP', 'MMDP')  
    CHM13_peak_data['Peak_name'] = CHM13_peak_data['Peak_name'].str.replace('LDAP', 'LMDP')  
    CHM13_peak_data['Integrated motif type'] = CHM13_peak_data['Integrated motif type'].str.replace('TCmix', 'TCnt')  
    CHM13_peak_data['Integrated motif type'] = CHM13_peak_data['Integrated motif type'].str.replace('AGmix', 'AGnt')  


    CHM13_HMDP = CHM13_peak_data[CHM13_peak_data['Peak_name'].str.contains('HMDP')] 


    G38_peak_data = pd.read_csv(infile + "G38-"+chr_parm+"_LP+_count.csv", encoding=get_encoding(infile +  "G38-"+chr_parm+"_LP+_count.csv"))

    G38_peak_data['Peak_name'] = G38_peak_data['Peak_name'].str.replace('G38', 'GRCh38')  
    G38_peak_data['Peak_name'] = G38_peak_data['Peak_name'].str.replace('HDAP', 'HMDP')  
    G38_peak_data['Peak_name'] = G38_peak_data['Peak_name'].str.replace('MDAP', 'MMDP')  
    G38_peak_data['Peak_name'] = G38_peak_data['Peak_name'].str.replace('LDAP', 'LMDP')  
    G38_peak_data['Integrated motif type'] = G38_peak_data['Integrated motif type'].str.replace('TCmix', 'TCnt')  
    G38_peak_data['Integrated motif type'] = G38_peak_data['Integrated motif type'].str.replace('AGmix', 'AGnt')  





    G38_gaps_data = pd.read_csv(gap_file + gap_dataf[0], encoding=get_encoding(gap_file + gap_dataf[0]))


    no_gap = 1
    chr_no = 1
    before_version = None

    # Add a new column for DataFrame
    G38_gaps_data['gap_no'] = None

    for index, row in G38_gaps_data.iterrows():
        gap_version = row[0]
        if no_gap == 1:
            chr_no_i = str(chr_no).zfill(2)
            no_gap_i = str(no_gap).zfill(3)
            gap_no = 'GRCh38' + '-' + chr_no_i + '-' + no_gap_i
            no_gap += 1
            before_version = gap_version
        else:
            if before_version == gap_version:
                no_gap_i = str(no_gap).zfill(3)
                gap_no = 'GRCh38' + '-' + chr_no_i + '-' + no_gap_i
                no_gap += 1
                before_version = gap_version
            else:
                no_gap = 1
                chr_no += 1
                if chr_no == 23:
                    chr_no_i = 'X'
                elif chr_no == 24:
                    chr_no_i = 'Y'
                if chr_no != 23 and chr_no != 24:
                    chr_no_i = str(chr_no).zfill(2)
                no_gap_i = str(no_gap).zfill(3)
                gap_no = 'GRCh38' + '-' + chr_no_i + '-' + no_gap_i
                no_gap += 1
                before_version = gap_version

        G38_gaps_data.at[index, 'gap_no'] = gap_no


    filtered_G38_gaps = G38_gaps_data[G38_gaps_data['gap_no'].str.contains('GRCh38-'+chr_parm)]

    deal_mummer_data['Peak_name_1'] = ''
    deal_mummer_data['Start_1'] = None
    deal_mummer_data['End_1'] = None
    deal_mummer_data['Motif type_1'] = ''
    deal_mummer_data['Representative main motif type_1'] =''
    deal_mummer_data['pD1RD sum_1'] = None
    deal_mummer_data['flag_copy_1'] = ''


    deal_mummer_data['Peak_name_2'] = ''
    deal_mummer_data['Start_2'] = None
    deal_mummer_data['End_2'] = None
    deal_mummer_data['Motif type_2'] = ''
    deal_mummer_data['Representative main motif type_2'] =''
    deal_mummer_data['pD1RD sum_2'] = None
    deal_mummer_data['flag_copy_2'] = ''


    #Detect String
    def check_char_in_string(char, string):
        split_string = string.split('/')
        return char in split_string[:3] 


    for index, row in CHM13_HMDP.iterrows():
        start_1 = row['S_position']
        end_1 = row['E_position']
        peak_1 = row['Peak_name']
        Motif_type_1 = row['Motif type']
        main_motif_type_1 = row['Integrated motif type']
        pD1RD_sum_1 = row['pD1RD sum']
        intersect_f =((deal_mummer_data['[S1]'] <= start_1) & (deal_mummer_data['[E1]'] >= end_1))
        if intersect_f.any():
            range_intersect = deal_mummer_data.loc[intersect_f]
            not_inserted = range_intersect['Extra_Column'] != 'Inserted'
            final_intersect = pd.Series(False, index=deal_mummer_data.index)
            final_intersect[not_inserted.index] = not_inserted.values
            if final_intersect.any():
                intersect = final_intersect
            else:
                intersect_l = (
                                      (deal_mummer_data['[S1]'] <= end_1) & (deal_mummer_data['[E1]'] >= start_1)
                              ) | (
                                      (deal_mummer_data['[S1]'] >= start_1) & (deal_mummer_data['[S1]'] <= end_1)
                              ) | (
                                      (deal_mummer_data['[S1]'] <= start_1) & (deal_mummer_data['[E1]'] >= end_1)
                              ) | (
                                      (deal_mummer_data['[E1]'] >= start_1) & (deal_mummer_data['[E1]'] <= end_1)
                              ) | (
                                      (deal_mummer_data['[S1]'] >= start_1) & (deal_mummer_data['[E1]'] <= end_1)
                              )
                intersect = intersect_l
        else:
            intersect_l = (
                (deal_mummer_data['[S1]'] <= end_1) & (deal_mummer_data['[E1]'] >= start_1)
            ) | (
                (deal_mummer_data['[S1]'] >= start_1) & (deal_mummer_data['[S1]'] <= end_1)
            ) | (
                (deal_mummer_data['[S1]'] <= start_1) & (deal_mummer_data['[E1]'] >= end_1)
            ) | (
                (deal_mummer_data['[E1]'] >= start_1) & (deal_mummer_data['[E1]'] <= end_1)
            ) | (
                (deal_mummer_data['[S1]'] >= start_1) & (deal_mummer_data['[E1]'] <= end_1)
            )
            intersect = intersect_l

        filtered_data = deal_mummer_data[intersect]
        duplicates = filtered_data.duplicated(subset=filtered_data.columns[:4], keep='last')
        df_no_duplicates = filtered_data[~duplicates]
        matching_rows = df_no_duplicates.index
        if len(matching_rows) > 0:
            if len(matching_rows) == 1:
                for match_index in matching_rows:
                    matching_row = deal_mummer_data.loc[match_index]
                    if matching_row['Peak_name_1'] == '':
                        deal_mummer_data.at[match_index, 'Peak_name_1'] = peak_1
                        deal_mummer_data.at[match_index, 'Start_1'] = start_1
                        deal_mummer_data.at[match_index, 'End_1'] = end_1
                        deal_mummer_data.at[match_index, 'Motif type_1'] = Motif_type_1
                        deal_mummer_data.at[match_index, 'Representative main motif type_1'] = main_motif_type_1
                        deal_mummer_data.at[match_index, 'pD1RD sum_1'] = pD1RD_sum_1
                    else:
                        if deal_mummer_data.loc[match_index+1]['flag_copy_1'] != 'copy_row_1':
                            new_row = matching_row.copy()
                            new_row['Peak_name_1'] = peak_1
                            new_row['Start_1'] = start_1
                            new_row['End_1'] = end_1
                            new_row['Motif type_1'] = Motif_type_1
                            new_row['Representative main motif type_1'] = main_motif_type_1
                            new_row['pD1RD sum_1'] = pD1RD_sum_1
                            new_row['flag_copy_1'] = 'copy_row_1'
                            deal_mummer_data = df_insert(deal_mummer_data, match_index+1, new_row)
                        else:
                            print("If it is not the same row, insert a row below it and copy the data from that row to match the numbered data")
                            for i in range(len(matching_rows)):
                                if deal_mummer_data.loc[match_index+i+2]['flag_copy_1'] != 'copy_row_1':
                                    new_row = deal_mummer_data.loc[match_index+i+1].copy()
                                    new_row['Peak_name_1'] = peak_1
                                    new_row['Start_1'] = start_1
                                    new_row['End_1'] = end_1
                                    new_row['flag_copy_1'] = 'copy_row_1'
                                    #df_new = insert(CHM13_HMDP, 1, df_add)
                                    deal_mummer_data = df_insert(deal_mummer_data, match_index+i+2, new_row)
                                    break
                            #print(CHM13_HMDP)
                    break
            else:
                j = 0
                for match_index in matching_rows:
                    matching_row = deal_mummer_data.loc[match_index+j]

                    if matching_row['Peak_name_1'] == '':

                        deal_mummer_data.at[match_index+j, 'Peak_name_1'] = peak_1
                        deal_mummer_data.at[match_index+j, 'Start_1'] = start_1
                        deal_mummer_data.at[match_index+j, 'End_1'] = end_1
                        deal_mummer_data.at[match_index+j, 'Motif type_1'] = Motif_type_1
                        deal_mummer_data.at[match_index+j, 'Representative main motif type_1'] = main_motif_type_1
                        deal_mummer_data.at[match_index+j, 'pD1RD sum_1'] = pD1RD_sum_1
                    else:
                        if deal_mummer_data.loc[match_index+j+1]['flag_copy_1'] != 'copy_row_1':

                            new_row = matching_row.copy()
                            new_row['Peak_name_1'] = peak_1
                            new_row['Start_1'] = start_1
                            new_row['End_1'] = end_1
                            new_row['Motif type_1'] = Motif_type_1
                            new_row['Representative main motif type_1'] = main_motif_type_1
                            new_row['pD1RD sum_1'] = pD1RD_sum_1
                            new_row['flag_copy_1'] = 'copy_row_1'
                            #df_new = insert(CHM13_HMDP, 1, df_add)
                            deal_mummer_data = df_insert(deal_mummer_data, match_index+j+1, new_row)
                            j+= 1
                        else:
                            print("If it is not the same row, insert a row below it and copy the data from that row to match the numbered data")

                            for i in range(len(matching_rows)):
                                if deal_mummer_data.loc[match_index+j+i+2]['flag_copy_1'] != 'copy_row_1':
                                    new_row = deal_mummer_data.loc[match_index+j+i+1].copy()
                                    new_row['Peak_name_1'] = peak_1
                                    new_row['Start_1'] = start_1
                                    new_row['End_1'] = end_1
                                    new_row['flag_copy_1'] = 'copy_row_1'
                                    #df_new = insert(CHM13_HMDP, 1, df_add)
                                    deal_mummer_data = df_insert(deal_mummer_data, match_index+j+i+2, new_row)
                                    j+=1
                                    break
        else:

            max_value = deal_mummer_data['[E1]'].max()

            if start_1 > max_value:
                match_index = len(deal_mummer_data)
                deal_mummer_data.at[match_index, 'Peak_name_1'] = peak_1
                deal_mummer_data.at[match_index, 'Start_1'] = start_1
                deal_mummer_data.at[match_index, 'End_1'] = end_1
                deal_mummer_data.at[match_index, 'Motif type_1'] = Motif_type_1
                deal_mummer_data.at[match_index, 'Representative main motif type_1'] = main_motif_type_1
                deal_mummer_data.at[match_index, 'pD1RD sum_1'] = pD1RD_sum_1
                deal_mummer_data.at[match_index,'Extra_Column'] = 'Inserted_last' 
            else:        
                print("\n"+"----------------------------------------------------------------")
                print(f'{peak_1}'+" :No matching interval data exists - please check the data")
                print("----------------------------------------------------------------"+"\n")
                pass
        #deal_mummer_data.to_csv(infile + "c13_g38_chr" + chr_parm + "_comparison.csv", index=False)
    deal_mummer_data.to_csv(infile+"c13_g38_chr"+chr_parm+"_comparison.csv", index=False)


    def calculate_intersection_length(start1, end1, start2, end2):

        start = max(start1, start2)

        end = min(end1, end2)


        if start > end:
            return 0


        length = end - start + 1

        return length


    deal_mummer_data_scan = deal_mummer_data[deal_mummer_data['Start_1'].notnull()]
    deal_mummer_data_scan.to_csv(infile+"c13_g38_chr"+chr_parm+"_comparison_1.csv", index=False)

    #print(deal_mummer_data_scan)


    list_deal_peak = [] 
    dict_deal_peak = {} 


    def check_character_and_df(character: str, df: pd.DataFrame) -> bool:
        def count_m(string: str) -> int:
            if 'm' not in string:
                return 1
            else:
                index_m = string.index('m')
                return int(string[index_m + 1:])
        
        input_character_count = count_m(character)
        column_sum = df['Peak_name_2'].apply(count_m).sum()
        
        return input_character_count == column_sum


    def check_character_and_df_r(character: str, df: pd.DataFrame) -> bool:
        def count_m(string: str) -> int:
            if 'm' not in string:
                return 1
            else:
                index_m = string.index('m')
                return int(string[index_m + 1:])

        # Return True if the length of the DataFrame is 1
        if len(df) == 1:
            return True

        input_character_count = count_m(character)
        column_sum = df['Peak_name_2'].apply(count_m).sum()

        return input_character_count == column_sum



    from typing import Tuple

    def check_character_and_df_bin(character: str, df: pd.DataFrame) -> Tuple[bool, pd.DataFrame]:
        def count_m(string: str) -> int:
            if 'm' not in string:
                return 1
            else:
                index_m = string.index('m')
                return int(string[index_m + 1:])

        input_character_count = count_m(character)
        df['count_m'] = df['Peak_name_2'].apply(count_m)
        column_sum = df['count_m'].sum()

        if column_sum > input_character_count:
            if len(df['Peak_name_2']) == 1:
                modified_df = df.drop(columns=['count_m'])
                return True, modified_df
            else:
                equal_count_rows = df[df['count_m'] == input_character_count]
                if len(equal_count_rows) == 1:
                    modified_df = equal_count_rows.drop(columns=['count_m'])
                    return True, modified_df
                else:
                    modified_df = df.drop(columns=['count_m'])
                    return False, modified_df
        else:
            modified_df = df.drop(columns=['count_m'])
            return True, modified_df


    def check_index_no(df: pd.DataFrame) -> bool:
        unique_values = df['index_no'].dropna().unique()
        if len(unique_values) > 1:
            return False
        else:
            return True


    from typing import List, Tuple

    def count_m(string: str) -> int:
        if 'm' not in string:
            return 1
        else:
            index_m = string.index('m')
            return int(string[index_m + 1:])


    def intersect_length(start_2: int, end_2: int, mummer_2_s_modify: int, mummer_2_e_modify: int) -> int:
        return max(0, min(end_2, mummer_2_e_modify) - max(start_2, mummer_2_s_modify) + 1)


    def combine_peaks(peak_1: str, df: pd.DataFrame, mummer_2_s_modify: int, mummer_2_e_modify: int) -> pd.DataFrame:
        input_count = count_m(peak_1)
        df_counts = [count_m(peak_name) for peak_name in df['Peak_name_2']]

        max_intersect = 0
        best_start_idx = None
        best_end_idx = None

        for start_idx in range(len(df_counts)):
            end_idx = start_idx
            current_count = 0
            while end_idx < len(df_counts) and current_count < input_count:
                current_count += df_counts[end_idx]
                end_idx += 1

            if current_count != input_count:
                continue

            start_2 = df.iloc[start_idx]['Start_2']
            end_2 = df.iloc[end_idx - 1]['End_2']

            intersect = intersect_length(start_2, end_2, mummer_2_s_modify, mummer_2_e_modify)

            if intersect > max_intersect:
                max_intersect = intersect
                best_start_idx = start_idx
                best_end_idx = end_idx - 1

        if best_start_idx is not None and best_end_idx is not None:
            best_df = df.iloc[best_start_idx:best_end_idx + 1].reset_index(drop=True)
            return best_df
        else:
            return pd.DataFrame()
            
    def combine_peaks_v2(peak_1: str, df: pd.DataFrame, mummer_2_s_modify: int, mummer_2_e_modify: int) -> pd.DataFrame:
        input_count = count_m(peak_1)
        df_counts = [count_m(peak_name) for peak_name in df['Peak_name_2']]

        max_intersect = 0
        best_start_idx = None
        best_end_idx = None
        found_equal_count = False

        for start_idx in range(len(df_counts)):
            end_idx = start_idx
            current_count = 0
            while end_idx < len(df_counts) and current_count < input_count:
                current_count += df_counts[end_idx]
                end_idx += 1

            if current_count == input_count:
                found_equal_count = True
                start_2 = df.iloc[start_idx]['Start_2']
                end_2 = df.iloc[end_idx - 1]['End_2']

                intersect = intersect_length(start_2, end_2, mummer_2_s_modify, mummer_2_e_modify)

                if intersect > max_intersect:
                    max_intersect = intersect
                    best_start_idx = start_idx
                    best_end_idx = end_idx - 1

        if found_equal_count:
            best_df = df.iloc[best_start_idx:best_end_idx + 1].reset_index(drop=True)
            return best_df
        else:
            return df


    def combine_peaks_v5(peak_1: str, df: pd.DataFrame, mummer_2_s_modify: int, mummer_2_e_modify: int) -> pd.DataFrame:
        input_count = count_m(peak_1)
        df_counts = [count_m(peak_name) for peak_name in df['Peak_name_2']]

        max_intersect = 0
        best_start_idx = None
        best_end_idx = None
        found_equal = False

        # Step 1: Find the best continuous subsequence with equal or maximum intersect
        for start_idx in range(len(df_counts)):
            end_idx = start_idx
            current_count = 0
            while end_idx < len(df_counts):
                current_count += df_counts[end_idx]
                if current_count >= input_count:
                    break
                end_idx += 1

            start_2 = df.iloc[start_idx]['Start_2']
            end_2 = df.iloc[min(end_idx, len(df) - 1)]['End_2']  

            intersect = intersect_length(start_2, end_2, mummer_2_s_modify, mummer_2_e_modify)

            if current_count == input_count:
                found_equal = True
                if intersect > max_intersect:
                    max_intersect = intersect
                    best_start_idx = start_idx
                    best_end_idx = end_idx
            elif not found_equal and intersect > max_intersect:
                max_intersect = intersect
                best_start_idx = start_idx
                best_end_idx = end_idx

        # Step 2: If no equal count subsequence is found, find the single peak with maximum intersect
        if not found_equal:
            max_intersect = 0
            best_single_idx = None
            for idx in range(len(df_counts)):
                start_2 = df.iloc[idx]['Start_2']
                end_2 = df.iloc[idx]['End_2']
                intersect = intersect_length(start_2, end_2, mummer_2_s_modify, mummer_2_e_modify)

                if intersect > max_intersect:
                    max_intersect = intersect
                    best_single_idx = idx

            if best_single_idx is not None:
                return df.iloc[[best_single_idx]].reset_index(drop=True)

        if best_start_idx is not None and best_end_idx is not None:
            best_df = df.iloc[best_start_idx:best_end_idx + 1].reset_index(drop=True)
            return best_df
        else:
            return pd.DataFrame()

            
            
    def extract_sequence(data):
        pattern = r'\[([A-Za-z]+)\]'
        match = re.search(pattern, data)

        if match:
            return match.group(1)
        else:
            return None

    def find_row_with_min_abs_difference(df, column_name, target_value, main_motif_1):
        main_motif_1 = extract_sequence(main_motif_1)

        df['extracted_main_motif_type_2'] = df['Representative main motif type_2'].apply(extract_sequence)

        equal_main_motif_1_rows = df['extracted_main_motif_type_2'] == main_motif_1
        filtered_df = df[equal_main_motif_1_rows]

        if len(filtered_df) == 1:
            return filtered_df.drop('extracted_main_motif_type_2', axis=1)

        min_abs_difference = (filtered_df[column_name] - target_value).abs().min()

        result_df = filtered_df[(filtered_df[column_name] - target_value).abs() == min_abs_difference]

        result_df = result_df.drop('extracted_main_motif_type_2', axis=1)

        return result_df

    for index_c, row_c in deal_mummer_data_scan.iterrows():
        if row_c['Peak_name_1'] != '':
            if row_c['Peak_name_1'] in list_deal_peak:
                dict_match_peak = {} #
                continue
            dict_match_peak = {}
            deal_peak_df = deal_mummer_data_scan[deal_mummer_data_scan['Peak_name_1'] == row_c['Peak_name_1']]
            list_deal_peak.append(row_c['Peak_name_1'])
            for index_d, row_d in deal_peak_df.iterrows():
                peak_1 = row_d['Peak_name_1']
                start_1 = row_d['Start_1']
                end_1 = row_d['End_1']
                peak_1_len = end_1 - start_1
                mummer_2_s = row_d['[S2]']
                mummer_2_e = row_d['[E2]']
                mummer_1_s = row_d['[S1]']
                mummer_1_e = row_d['[E1]']
                Extra_flag = row_d['Extra_Column']
                Motif_type_1 = row_d['Motif type_1']
                main_motif_type_1 = row_d['Representative main motif type_1']
                pD1RD_sum_1 = row_d['pD1RD sum_1']
                if 'm' in peak_1:
                    m_count = int(peak_1.split('m')[1])
                else:
                    m_count = 0
                diff_position,nearest_position = find_nearest_position(start_1, mummer_1_s, mummer_1_e)
                if mummer_2_e > mummer_2_s and mummer_1_e > mummer_1_s:
                    if nearest_position == 'start':
                        mummer_2_s_modify = mummer_2_s + diff_position-param_region_1
                        if mummer_2_s_modify > 0:
                            mummer_2_s_modify = mummer_2_s_modify
                        else:
                            print(f'{peak_1}'+"---"+" :Calculate the starting interval as a negative number-1")
                            mummer_2_s_modify = 1
                        mummer_2_e_modify = mummer_2_s_modify + peak_1_len+param_region_1
                    elif nearest_position == 'end':
                        #print(peak_1)
                        mummer_2_s_modify = mummer_2_e + diff_position - param_region_1
                        if mummer_2_s_modify > 0:
                            mummer_2_s_modify = mummer_2_s_modify
                        else:
                            mummer_2_s_modify = 1
                            print(f'{peak_1}'+"---"+" :Calculate the starting interval as a negative number-2")
                        mummer_2_e_modify = mummer_2_s_modify + peak_1_len + param_region_1
                    else:
                        print("Data error---1")
                else:
                    #print(f'{peak_1}' + "---" + f'{row_d["index_no"]}' + " :Interval reverse similarity")
                    if nearest_position == 'start':
                        mummer_2_s_modify = mummer_2_e + diff_position-param_region_1
                        if mummer_2_s_modify > 0:
                            mummer_2_s_modify = mummer_2_s_modify
                        else:
                            mummer_2_s_modify = 1
                        mummer_2_e_modify = mummer_2_s_modify + peak_1_len+param_region_1
                    elif nearest_position == 'end':
                        #print(peak_1)
                        mummer_2_s_modify = mummer_2_s + diff_position - param_region_1
                        if mummer_2_s_modify > 0:
                            mummer_2_s_modify = mummer_2_s_modify
                        else:
                            mummer_2_s_modify = 1
                        mummer_2_e_modify = mummer_2_s_modify + peak_1_len + param_region_1
                    else:
                        print("Data error---1")
                if mummer_2_s_modify > mummer_2_e_modify:
                    print(f'{peak_1}'+" ： Error in interval calculation, please check")
                intersect_data = (
                                    (G38_peak_data['S_position'] <= mummer_2_e_modify) & (G38_peak_data['E_position'] >= mummer_2_s_modify)
                            ) | (
                                    (G38_peak_data['S_position'] >= mummer_2_s_modify) & (G38_peak_data['S_position'] <= mummer_2_e_modify)
                            ) | (
                                    (G38_peak_data['S_position'] <= mummer_2_s_modify) & (G38_peak_data['E_position'] >= mummer_2_e_modify)
                            ) | (
                                    (G38_peak_data['E_position'] >= mummer_2_s_modify) & (G38_peak_data['E_position'] <= mummer_2_e_modify)
                            ) | (
                                    (G38_peak_data['S_position'] >= mummer_2_s_modify) & (G38_peak_data['E_position'] <= mummer_2_e_modify)
                            )
                matching_rows = G38_peak_data[intersect_data].index
                if len(matching_rows)>0:
                    for match_index in matching_rows:
                        matching_row = G38_peak_data.loc[match_index]
                        start_2 = matching_row['S_position']
                        end_2 = matching_row['E_position']
                        peak_2 = matching_row['Peak_name']
                        Motif_type_2 = matching_row['Motif type']
                        main_motif_type_2 = matching_row['Integrated motif type']
                        pD1RD_sum_2 = matching_row['pD1RD sum']
                        param1_cpd_1 = main_motif_type_1.split(']')[0].split('[')[1]
                        param2_cpd_2 = main_motif_type_2.split(']')[0].split('[')[1]
                        keys_values = [('Peak_name_2', peak_2), ('Start_2', start_2), ('End_2', end_2),
                                       ('Motif type_2', Motif_type_2),
                                       ('Representative main motif type_2', main_motif_type_2),
                                       ('pD1RD sum_2', pD1RD_sum_2)]
                        if param1_cpd_1 == param2_cpd_2 and 'mix' not in param1_cpd_1: 
                            if peak_2 not in dict_match_peak.get('Peak_name_2', []):
                                dict_match_peak.setdefault('Peak_name_2', []).append(peak_2)
                                dict_match_peak.setdefault('Start_2', []).append(start_2)
                                dict_match_peak.setdefault('End_2', []).append(end_2)
                                dict_match_peak.setdefault('Motif type_2', []).append(Motif_type_2)
                                dict_match_peak.setdefault('Representative main motif type_2', []).append(main_motif_type_2)
                                dict_match_peak.setdefault('pD1RD sum_2', []).append(pD1RD_sum_2)
                        else:
                            if 'nt' in param1_cpd_1 and 'nt' not in param2_cpd_2:
                                if (check_char_in_string(param2_cpd_2,Motif_type_1)):
                                    if peak_2 not in dict_match_peak.get('Peak_name_2', []):
                                        for key, value in keys_values:
                                            dict_match_peak.setdefault(key, []).append(value)
                            elif 'nt' not in param1_cpd_1 and 'nt' in param2_cpd_2:
                                if (check_char_in_string(param1_cpd_1,Motif_type_2)):
                                    if peak_2 not in dict_match_peak.get('Peak_name_2', []):
                                        for key, value in keys_values:
                                            dict_match_peak.setdefault(key, []).append(value)

            if len(dict_match_peak) >0:
                df_match_peak = pd.DataFrame.from_dict(dict_match_peak)
                #print(df_match_peak)
                filtered_peak_df = deal_mummer_data[deal_mummer_data['Peak_name_1'] == peak_1]
                filtered_peak_df_copy = filtered_peak_df.copy()
                filtered_peak_df_copy.loc[:, 'Extra_Column'] = filtered_peak_df_copy['Extra_Column'].fillna("")
                filtered_peak_df_flag = (~filtered_peak_df_copy['Extra_Column'].str.contains('Inserted')).all()
                if filtered_peak_df_flag: #Check if all are within a highly similar range
                    if (check_character_and_df_r(peak_1,df_match_peak)): #Check if peak is equal
                        pass
                    else:
                        df_match_peak_copy = df_match_peak
                        if 'm' not in peak_1:
                            df_match_peak_deal = find_row_with_min_abs_difference(df_match_peak, 'pD1RD sum_2', pD1RD_sum_1,main_motif_type_1) #处理高度相似区间不同bin的数据
                            #df_match_peak = df_match_peak_deal
                            if len(df_match_peak_deal) == 0:
                                sorted_df = df_match_peak.sort_values(by="Start_2", ascending=True)
                                combine_peaks_df = combine_peaks(peak_1, sorted_df, mummer_2_s_modify, mummer_2_e_modify)
                                df_match_peak = combine_peaks_df
                            else:
                                df_match_peak = df_match_peak_deal
                        else:
                            sorted_df = df_match_peak.sort_values(by="Start_2", ascending=True)
                            combine_peaks_df = combine_peaks_v2(peak_1,sorted_df,mummer_2_s_modify,mummer_2_e_modify)
                            df_match_peak = combine_peaks_df
                        # print(peak_1)
                        #print(df_match_peak)
                        # print(df_match_peak_copy)
                        if (check_character_and_df_r(peak_1,df_match_peak)): #Check if peak is equal
                            pass
                        else:
                            if len(df_match_peak) != 1:
                                print("\n"+"------------Data processing error-----------")
                                print(f'{peak_1}'+"---"+f'{df_match_peak}'+f'{df_match_peak_copy}'+" :Inconsistent bin in the same interval")
                                print("------------Data processing error-----------"+"\n")
                else: #There is cross regional data
                    result, modified_df = check_character_and_df_bin(peak_1, df_match_peak)
                    df_match_peak = modified_df

                    if (result): #Check if peak is equal
                        pass
                    else:
                        #Add check and delete excess matching data
                        filtered_peak_df_copy1 = filtered_peak_df.copy()
                        if (check_index_no(filtered_peak_df_copy1)):
                            sorted_df = df_match_peak.sort_values(by="Start_2", ascending=True)
                            combine_peaks_df = combine_peaks_v5(peak_1,sorted_df,mummer_2_s_modify,mummer_2_e_modify)
                            df_match_peak = combine_peaks_df
                        else: 
                            df_no_na_many_region = filtered_peak_df_copy1.dropna(subset=["index_no"])
                            filtered_df_many_region = df_no_na_many_region.drop_duplicates(subset=["index_no"], keep="first")
                            mummer_2_s = filtered_df_many_region.iloc[0]['[S2]']
                            mummer_2_e = filtered_df_many_region.iloc[0]['[E2]']
                            mummer_1_s = filtered_df_many_region.iloc[0]['[S1]']
                            mummer_1_e = filtered_df_many_region.iloc[0]['[E1]']
                            mummer_2_s_modify = mummer_2_e + filtered_df_many_region.iloc[0]['Start_1']-mummer_1_e
                            mummer_2_e_modify = mummer_2_s_modify + peak_1_len
                            sorted_df = df_match_peak.sort_values(by="Start_2", ascending=True)
                            combine_peaks_df = combine_peaks(peak_1,sorted_df,mummer_2_s_modify,mummer_2_e_modify)
                            df_match_peak = combine_peaks_df
                flag_one = 0
                for index_e, row_e in df_match_peak.iterrows():
                    if flag_one != 0:
                        break
                    start_2 = row_e['Start_2']
                    end_2 = row_e['End_2']
                    peak_2 = row_e['Peak_name_2']
                    Motif_type_2 = row_e['Motif type_2']
                    main_motif_type_2 = row_e['Representative main motif type_2']
                    pD1RD_sum_2 = row_e['pD1RD sum_2']
                    intersect_e_f =((deal_mummer_data['[S1]'] <= start_1) & (deal_mummer_data['[E1]'] >= end_1))
                    if intersect_e_f.any():
                        if (deal_mummer_data[intersect_e_f]['Extra_Column'] != 'Inserted').all() and filtered_peak_df_flag:
                            intersect_e = intersect_e_f
                            if 'm' not in peak_1: 
                                flag_one = 1
                        else:
                            intersect_e_l = (
                                                (deal_mummer_data['[S2]'] <= end_2) & (
                                                        deal_mummer_data['[S2]'] >= start_2)
                                            ) | (
                                                    (deal_mummer_data['[S2]'] >= start_2) & (
                                                        deal_mummer_data['[S2]'] <= end_2)
                                            ) | (
                                                    (deal_mummer_data['[S2]'] <= start_2) & (
                                                        deal_mummer_data['[E2]'] >= end_2)
                                            ) | (
                                                    (deal_mummer_data['[E2]'] >= start_2) & (
                                                        deal_mummer_data['[E2]'] <= end_2)
                                            ) | (
                                                    (deal_mummer_data['[S2]'] >= start_2) & (
                                                        deal_mummer_data['[E2]'] <= end_2)
                                            )
                            intersect_e = intersect_e_l
                    else:                
                        intersect_e_l = (
                            (deal_mummer_data['[S2]'] <= end_2) & (deal_mummer_data['[S2]'] >= start_2)
                        ) | (
                            (deal_mummer_data['[S2]'] >= start_2) & (deal_mummer_data['[S2]'] <= end_2)
                        ) | (
                            (deal_mummer_data['[S2]'] <= start_2) & (deal_mummer_data['[E2]'] >= end_2)
                        ) | (
                            (deal_mummer_data['[E2]'] >= start_2) & (deal_mummer_data['[E2]'] <= end_2)
                        ) | (
                            (deal_mummer_data['[S2]'] >= start_2) & (deal_mummer_data['[E2]'] <= end_2)
                        )
                        intersect_e = intersect_e_l
                    filtered_data_e = deal_mummer_data[intersect_e]
                    duplicates_e = filtered_data_e.duplicated(subset=filtered_data_e.columns[:11], keep='last')
                    df_no_duplicates_e = filtered_data_e[~duplicates_e]
                    filtered_df = df_no_duplicates_e[df_no_duplicates_e['Peak_name_1'] == peak_1]
                    matching_rows_e = filtered_df.index
                    #print(filtered_df)
                    if len(matching_rows_e) > 0:
                        if len(matching_rows_e) == 1:
                            for match_index_e in matching_rows_e:
                                matching_row = deal_mummer_data.loc[match_index_e]
                                if matching_row['Peak_name_2'] == '':
                                    deal_mummer_data.at[match_index_e, 'Peak_name_2'] = peak_2
                                    deal_mummer_data.at[match_index_e, 'Start_2'] = start_2
                                    deal_mummer_data.at[match_index_e, 'End_2'] = end_2
                                    deal_mummer_data.at[match_index_e, 'Motif type_2'] = Motif_type_2
                                    deal_mummer_data.at[match_index_e, 'Representative main motif type_2'] = main_motif_type_2
                                    deal_mummer_data.at[match_index_e, 'pD1RD sum_2'] = pD1RD_sum_2
                                else:
                                    if match_index_e+1 < len(deal_mummer_data):
                                        if deal_mummer_data.loc[match_index_e+1]['flag_copy_2'] != 'copy_row_2':
                                            new_row = matching_row.copy()
                                            new_row['Peak_name_2'] = peak_2
                                            new_row['Start_2'] = start_2
                                            new_row['End_2'] = end_2
                                            new_row['Motif type_2'] = Motif_type_2
                                            new_row['Representative main motif type_2'] = main_motif_type_2
                                            new_row['pD1RD sum_2'] = pD1RD_sum_2
                                            new_row['flag_copy_2'] = 'copy_row_2'
                                            #df_new = insert(CHM13_HMDP, 1, df_add)
                                            deal_mummer_data = df_insert(deal_mummer_data, match_index_e+1, new_row)
                                        else:
                                            print("If it is not the same row, insert a row below it and copy the data from that row to match the numbered data")
                                    else:
                                        new_row = matching_row.copy()
                                        new_row['Peak_name_2'] = peak_2
                                        new_row['Start_2'] = start_2
                                        new_row['End_2'] = end_2
                                        new_row['Motif type_2'] = Motif_type_2
                                        new_row['Representative main motif type_2'] = main_motif_type_2
                                        new_row['pD1RD sum_2'] = pD1RD_sum_2
                                        new_row['flag_copy_2'] = 'copy_row_2'
                                        deal_mummer_data = deal_mummer_data.append(new_row, ignore_index=True)
                                break
                        else:
                            j = 0
                            for match_index_e in matching_rows_e:
                                matching_row = deal_mummer_data.loc[match_index_e+j]
                                if matching_row['Peak_name_2'] == '':
                                    deal_mummer_data.at[match_index_e+j, 'Peak_name_2'] = peak_2
                                    deal_mummer_data.at[match_index_e+j, 'Start_2'] = start_2
                                    deal_mummer_data.at[match_index_e+j, 'End_2'] = end_2
                                    deal_mummer_data.at[match_index_e+j, 'Motif type_2'] = Motif_type_2
                                    deal_mummer_data.at[match_index_e+j, 'Representative main motif type_2'] = main_motif_type_2
                                    deal_mummer_data.at[match_index_e+j, 'pD1RD sum_2'] = pD1RD_sum_2
                                else:
                                    if deal_mummer_data.loc[match_index_e+j+1]['flag_copy_2'] != 'copy_row_2':
                                        new_row = matching_row.copy()
                                        new_row['Peak_name_2'] = peak_2
                                        new_row['Start_2'] = start_2
                                        new_row['End_2'] = end_2
                                        new_row['Motif type_2'] = Motif_type_2
                                        new_row['Representative main motif type_2'] = main_motif_type_2
                                        new_row['pD1RD sum_2'] = pD1RD_sum_2
                                        new_row['flag_copy_2'] = 'copy_row_2'
                                        #df_new = insert(CHM13_HMDP, 1, df_add)
                                        deal_mummer_data = df_insert(deal_mummer_data, match_index_e+j+1, new_row)
                                        j+= 1
                                    else:
                                        print("If it is not the same row, insert a row below it and copy the data from that row to match the numbered data")
                    else:
                        print("\n"+"----------------------------------------------------------------")
                        print(f'{peak_1}'+"---"+f'{peak_2}'+" :No matching interval data exists - please check the data")
                        print("----------------------------------------------------------------"+"\n")
                        pass
            else:
                intersect_gap = (
                                         (filtered_G38_gaps['start'] <= mummer_2_e_modify) & (
                                             filtered_G38_gaps['stop'] >= mummer_2_s_modify)
                                 ) | (
                                         (filtered_G38_gaps['start'] >= mummer_2_s_modify) & (
                                             filtered_G38_gaps['start'] <= mummer_2_e_modify)
                                 ) | (
                                         (filtered_G38_gaps['start'] <= mummer_2_s_modify) & (
                                             filtered_G38_gaps['stop'] >= mummer_2_e_modify)
                                 ) | (
                                         (filtered_G38_gaps['stop'] >= mummer_2_s_modify) & (
                                             filtered_G38_gaps['stop'] <= mummer_2_e_modify)
                                 ) | (
                                         (filtered_G38_gaps['start'] >= mummer_2_s_modify) & (
                                             filtered_G38_gaps['stop'] <= mummer_2_e_modify)
                                 )
                matching_gap_rows = filtered_G38_gaps[intersect_gap].index
                #print("--------------------Check the data comparison of motif--------------------------")
                if len(matching_gap_rows) != 0:
                    #print(f'{peak_1}' + "------" + "GAP")
                    deal_mummer_data.loc[(deal_mummer_data['Peak_name_1'] == peak_1) & (deal_mummer_data['Extra_Column'] != 'Inserted_last'), 'Peak_name_2'] = 'GAP'
                    deal_mummer_data.loc[(deal_mummer_data['Peak_name_1'] == peak_1) & (deal_mummer_data['Extra_Column'] != 'Inserted_last'), 'pD1RD sum_2'] = '-'
                else: 
                    filtered_data = deal_mummer_data[deal_mummer_data['Extra_Column'] != 'Inserted']
                    prev_row_index = None
                    next_row_index = None

                    for index, row in filtered_data.iterrows():
                        if row['[S1]'] > start_1:
                            next_row_index = index
                            break
                        prev_row_index = index

                    found_adjacent_rows = False
                    if prev_row_index is not None and next_row_index is not None:
                        prev_row = filtered_data.loc[prev_row_index]
                        next_row = filtered_data.loc[next_row_index]
                        if prev_row['index_no'] + 1 == next_row['index_no']:
                            found_adjacent_rows = True
                    if found_adjacent_rows:
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore", category=FutureWarning)
                            result = pd.DataFrame()
                            result = result.append(prev_row)
                            result = result.append(next_row)
                    else:
                        result = pd.DataFrame()
                    if len(result) != 0:
                        for index_d, row_d in result.iterrows():
                            mummer_2_s = row_d['[S2]']
                            mummer_2_e = row_d['[E2]']
                            mummer_1_s = row_d['[S1]']
                            mummer_1_e = row_d['[E1]']
                            diff_position,nearest_position = find_nearest_position(start_1, mummer_1_s, mummer_1_e)
                            if mummer_2_e > mummer_2_s and mummer_1_e > mummer_1_s:
                                if nearest_position == 'start':
                                    mummer_2_s_modify = mummer_2_s + diff_position-param_region_1
                                    if mummer_2_s_modify > 0:
                                        mummer_2_s_modify = mummer_2_s_modify
                                    else:
                                        print(f'{peak_1}'+"---"+" :Calculate the starting interval as a negative number-1")
                                        mummer_2_s_modify = 1
                                    mummer_2_e_modify = mummer_2_s_modify + peak_1_len+param_region_1
                                elif nearest_position == 'end':
                                    #print(peak_1)
                                    mummer_2_s_modify = mummer_2_e + diff_position - param_region_1
                                    if mummer_2_s_modify > 0:
                                        mummer_2_s_modify = mummer_2_s_modify
                                    else:
                                        mummer_2_s_modify = 1
                                        print(f'{peak_1}'+"---"+" :Calculate the starting interval as a negative number-2")
                                    mummer_2_e_modify = mummer_2_s_modify + peak_1_len + param_region_1
                                else:
                                    print("Data error---1")
                            else:
                                if nearest_position == 'start':
                                    mummer_2_s_modify = mummer_2_e + diff_position-param_region_1
                                    if mummer_2_s_modify > 0:
                                        mummer_2_s_modify = mummer_2_s_modify
                                    else:
                                        mummer_2_s_modify = 1
                                    mummer_2_e_modify = mummer_2_s_modify + peak_1_len+param_region_1
                                elif nearest_position == 'end':
                                    #print(peak_1)
                                    mummer_2_s_modify = mummer_2_s + diff_position - param_region_1
                                    if mummer_2_s_modify > 0:
                                        mummer_2_s_modify = mummer_2_s_modify
                                    else:
                                        mummer_2_s_modify = 1
                                    mummer_2_e_modify = mummer_2_s_modify + peak_1_len + param_region_1
                                else:
                                    print("Data error---1")
                            if mummer_2_s_modify > mummer_2_e_modify:
                                print(f'{peak_1}'+" ： Error in interval calculation, please check")
                            intersect_gap = (
                                             (filtered_G38_gaps['start'] <= mummer_2_e_modify) & (
                                                 filtered_G38_gaps['stop'] >= mummer_2_s_modify)
                                     ) | (
                                             (filtered_G38_gaps['start'] >= mummer_2_s_modify) & (
                                                 filtered_G38_gaps['start'] <= mummer_2_e_modify)
                                     ) | (
                                             (filtered_G38_gaps['start'] <= mummer_2_s_modify) & (
                                                 filtered_G38_gaps['stop'] >= mummer_2_e_modify)
                                     ) | (
                                             (filtered_G38_gaps['stop'] >= mummer_2_s_modify) & (
                                                 filtered_G38_gaps['stop'] <= mummer_2_e_modify)
                                     ) | (
                                             (filtered_G38_gaps['start'] >= mummer_2_s_modify) & (
                                                 filtered_G38_gaps['stop'] <= mummer_2_e_modify)
                                     )
                            matching_gap_rows_1 = filtered_G38_gaps[intersect_gap].index
                            #print("--------------------Check the data comparison of motif--------------------------")
                            if len(matching_gap_rows_1) != 0:
                                deal_mummer_data.loc[(deal_mummer_data['Peak_name_1'] == peak_1) & (deal_mummer_data['Extra_Column'] != 'Inserted_last'), 'Peak_name_2'] = 'GAP'
                                deal_mummer_data.loc[(deal_mummer_data['Peak_name_1'] == peak_1) & (deal_mummer_data['Extra_Column'] != 'Inserted_last'), 'pD1RD sum_2'] = '-'
                            else:
                                deal_mummer_data.loc[deal_mummer_data['Peak_name_1'] == peak_1, 'Peak_name_2'] = 'No peak'
                                deal_mummer_data.loc[deal_mummer_data['Peak_name_1'] == peak_1, 'pD1RD sum_2'] = '-'
                    else:
                        deal_mummer_data.loc[deal_mummer_data['Peak_name_1'] == peak_1, 'Peak_name_2'] = 'No peak'
                        deal_mummer_data.loc[deal_mummer_data['Peak_name_1'] == peak_1, 'pD1RD sum_2'] = '-'
    deal_mummer_data.to_csv(infile+"c13_g38_chr"+chr_parm+"_comparison_2.csv", index=False)
    
    
    def is_empty_c(value):
        return value is None or value == "" or pd.isnull(value)


    def combine_columns(row, col1, col2):
        if not is_empty_c(row[col1]) and not is_empty_c(row[col2]):
            return f"{row[col1]}({row[col2]})"
        elif not is_empty_c(row[col1]):
            return row[col1]
        else:
            return None

    deal_mummer_data['ucsc_name_1'] = deal_mummer_data.apply(lambda row: combine_columns(row, 'Peak_name_1', 'Representative main motif type_1'), axis=1)
    deal_mummer_data['ucsc_name_2'] = deal_mummer_data.apply(lambda row: combine_columns(row, 'Peak_name_2', 'Representative main motif type_2'), axis=1)
    
    
    def create_dict_from_dataframe(df):
        result_dict = {}
        for _, row in df.iterrows():
            key = row['Peak_name_1']
            value = row['Peak_name_2']

            if isinstance(key, str):
                key = key.strip()
            if isinstance(value, str):
                value = value.strip()

            if (not key or key == ""):
                continue

            if key not in result_dict:
                result_dict[key] = [value]  
            else:

                if not any(result_dict[key]) and value:
                    result_dict[key] = [value]

                elif value and value not in result_dict[key]:
                    result_dict[key].append(value)

        return result_dict
        
    dict_check = create_dict_from_dataframe(deal_mummer_data)

    for key, values in dict_check.items():
        for value in values:
            if value == "" or pd.isna(value):
                print(f"{key}: {value}")



    def create_dict_no_peak(df, key_column, peak_alt_column):
        temp_dict = {}
        result_dict = {}


        for _, row in df.iterrows():
            key = row[key_column]

            if isinstance(key, str):
                key = key.strip()

            if not key or key == "":
                continue

            peak_alt_value = row[peak_alt_column]

            if peak_alt_value == "H-No peak":
                result_dict[key] =  peak_alt_value
                continue
        return result_dict


    def create_dict_diff(df, key_column, value_columns, check_column,alt_value_column=None):
        temp_dict = {}
        result_dict = {}


        for _, row in df.iterrows():
            key = row[key_column]
            h_match_type = row[check_column]
            if isinstance(key, str):
                key = key.strip()

            if not key or key == "":
                continue

            alt_value = row[alt_value_column]

            if alt_value == "GAP" or alt_value == "No peak":
                result_dict[key] = "-"
                continue
            if h_match_type and ('H' in h_match_type or 'M' in h_match_type or 'L' in h_match_type):
                if h_match_type.split('-')[0] != h_match_type.split('-')[1]:
                    result_dict[key] = "-"
                    continue
            if key not in temp_dict:
                temp_dict[key] = {
                    "value_1": row[value_columns[0]],
                    "value_2_list": [row[value_columns[1]]],
                }
            else:
                if row[value_columns[1]] not in temp_dict[key]["value_2_list"]:
                    temp_dict[key]["value_2_list"].append(row[value_columns[1]])


        for key, values in temp_dict.items():
            value_1 = values["value_1"]
            value_2_list = values["value_2_list"]

            if isinstance(value_1, str):
                value_1 = value_1.strip()

            try:
                value_1 = float(value_1)
            except ValueError:
                continue

            sum_value_2 = 0

            for value_2 in value_2_list:
                if isinstance(value_2, str):
                    value_2 = value_2.strip()

                if value_2 is None:
                    continue

                try:
                    value_2 = float(value_2)
                except ValueError:
                    continue

                sum_value_2 += value_2

            diff = abs(value_1 - sum_value_2)
            result_dict[key] = diff

        return result_dict




    def fill_absolute_difference(row, key_column, result_dict):
        key = row[key_column]

        if isinstance(key, str):
            key = key.strip()

        if not key or key == "":
            return None
        
        if key in result_dict:
            return result_dict[key]
        else:
            return None


    from collections import defaultdict
    def create_dict_match_type(df, key_column, peak_alt_column):
        temp_dict = {}
        result_dict = {}


        for _, row in df.iterrows():
            key = row[key_column]

            if isinstance(key, str):
                key = key.strip()

            if not key or key == "":
                continue

            peak_alt_value = row[peak_alt_column]

            if peak_alt_value == "GAP" or peak_alt_value == "No peak":
                result_dict[key] =  f"{peak_alt_value}"
                continue

            if not peak_alt_value or not isinstance(peak_alt_value, str) or peak_alt_value.strip() == "":
                continue

            if key not in temp_dict:
                temp_dict[key] = [peak_alt_value]
            else:
                if peak_alt_value not in temp_dict[key]:
                    temp_dict[key].append(peak_alt_value)

        for key, peak_names in temp_dict.items():
            counts = defaultdict(int)

            for peak_name in peak_names:

                peak_name = peak_name.split('-')[-1]


                match = re.search(r'(H|M|L)MDP(\d+)m(\d+)', peak_name)
                if match:
                    peak_type = match.group(1)
                    count = int(match.group(3))
                else:

                    match = re.search(r'(H|M|L)MDP(\d+)', peak_name)
                    if match:
                        peak_type = match.group(1)
                        count = 1
                    else:
                        continue

                counts[peak_type] += count


            key_m_matches = re.findall(r'm(\d+)', key)
            if key_m_matches:
                key_m_count = int(key_m_matches[0])
            else:
                key_m_count = 1

            result = []
            for peak_type, count in counts.items():
                result.append(f"{peak_type}{count}")

            if not result:
                result.append("NONE")
            result_type = None
            priority = {"H": 1, "M": 2, "L": 3}

            for item in result:
                peak_type = item[0]
                if result_type is None or priority[peak_type] < priority[result_type]:
                    result_type = peak_type

            #result_dict[key] = f"H{key_m_count}-{'/'.join(result)}"
            result_dict[key] = f"HMDP-{result_type}MDP"

        return result_dict
        


    def create_dict_match_type_j(df, key_column, peak_alt_column):
        temp_dict = {}
        result_dict = {}


        for _, row in df.iterrows():
            key = row[key_column]

            if isinstance(key, str):
                key = key.strip()

            if not key or key == "":
                continue

            peak_alt_value = row[peak_alt_column]

            if peak_alt_value == "GAP" or peak_alt_value == "No peak":
                result_dict[key] = peak_alt_value
                continue

            if not peak_alt_value or not isinstance(peak_alt_value, str) or peak_alt_value.strip() == "":
                continue

            if key not in temp_dict:
                temp_dict[key] = [peak_alt_value]
            else:
                if peak_alt_value not in temp_dict[key]:
                    temp_dict[key].append(peak_alt_value)

        for key, peak_names in temp_dict.items():
            counts = defaultdict(int)

            for peak_name in peak_names:

                peak_name = peak_name.split('-')[-1]


                match = re.search(r'(H|M|L)MDP(\d+)m(\d+)', peak_name)
                if match:
                    peak_type = match.group(1)
                    count = int(match.group(3))
                else:

                    match = re.search(r'(H|M|L)MDP(\d+)', peak_name)
                    if match:
                        peak_type = match.group(1)
                        count = 1
                    else:
                        continue

                counts[peak_type] += count


            key_m_matches = re.findall(r'm(\d+)', key)
            if key_m_matches:
                key_m_count = int(key_m_matches[0])
            else:
                key_m_count = 1

            result = []
            for peak_type, count in counts.items():
                result.append(f"{peak_type}{count}")

            if not result:
                result.append("NONE")

            result_dict[key] = f"H{key_m_count}-{'/'.join(result)}"

        return result_dict        
        
        
                    
    def create_dict_match_level(df, key_column, peak_alt_column):
        temp_dict = {}
        for _, row in df.iterrows():
            key = row[key_column]

            if isinstance(key, str):
                key = key.strip()

            if not key or key == "":
                continue
            peak_alt_value = row[peak_alt_column]
            if isinstance(peak_alt_value, (int, float)):
                peak_alt_value = str(peak_alt_value)
            if peak_alt_value is None or not isinstance(peak_alt_value, str):
                continue
            peak_alt_value = peak_alt_value.strip()
            if peak_alt_value == "":
                continue
            if key not in temp_dict:
                temp_dict[key] = [peak_alt_value]
            else:
                if peak_alt_value not in temp_dict[key]:
                    temp_dict[key].append(peak_alt_value)

        result_dict = {}
        #print(temp_dict)
        for key, values in temp_dict.items():
            match_level_found = False
            for value in values:
                if value == "GAP" or value == "No peak":
                    result_dict[key] =  "-"
                    match_level_found = True
                    break
                else:
                    try:
                        diff = float(value)
                    except ValueError:
                        continue

                    if diff >= 0 and diff <= 30:
                        result_dict[key] = "0-30"
                    elif diff >= 31 and diff <= 60:
                        result_dict[key] = "31-60"
                    elif diff >= 61 and diff <= 90:
                        result_dict[key] = "61-90"
                    elif diff >= 91 and diff <= 120:
                        result_dict[key] = "91-120"
                    elif diff >= 121 and diff <= 150:
                        result_dict[key] = "121-150"
                    elif diff >= 151 and diff <= 180:
                        result_dict[key] = "151-180"
                    elif diff >= 181 and diff <= 210:
                        result_dict[key] = "181-210"
                    elif diff >= 211 and diff <= 240:
                        result_dict[key] = "211-240"
                    elif diff >= 241 and diff <= 270:
                        result_dict[key] = "241-270"
                    elif diff >= 271 and diff <= 300:
                        result_dict[key] = "271-300"                        
                    else:
                        result_dict[key] = "301-"
                if result_dict[key]:
                    match_level_found = True
                    break
            if not match_level_found:
                result_dict[key] = "-"

        return result_dict
        
    def create_dict_match_level_many(df, key_column, peak_alt_column):
        temp_dict = {}
        for _, row in df.iterrows():
            key = row[key_column]
            #print(row[peak_alt_column])
            if isinstance(key, str):
                key = key.strip()

            if not key or key == "":
                continue
            peak_alt_value = row[peak_alt_column]
            if isinstance(peak_alt_value, (int, float)):
                peak_alt_value = str(peak_alt_value)
            if peak_alt_value is None or not isinstance(peak_alt_value, str):
                continue
            peak_alt_value = peak_alt_value.strip()
            if peak_alt_value == "":
                continue
            if key not in temp_dict:
                temp_dict[key] = [peak_alt_value]
            else:
                if peak_alt_value not in temp_dict[key]:
                    temp_dict[key].append(peak_alt_value)

        result_dict_1 = {}
        result_dict_2 = {}
        result_dict_3 = {}
        #print(temp_dict)
        for key, values in temp_dict.items():
            match_level_found = False
            for value in values:
                if value == "GAP" or value == "No peak":
                    #result_dict[key] =  f"H-{value}"
                    match_level_found = True
                    break
                else:
                    try:
                        diff = float(value)
                    except ValueError:
                        continue

                    if diff >= 0 and diff <= 30:
                        result_dict_1[key] = "0-30"
                    else:
                        result_dict_1[key] = ""
                    if diff >= 0 and diff <= 20:
                        result_dict_2[key] = "0-20"
                    if diff >= 0 and diff <= 10:
                        result_dict_3[key] = "0-10"
                if result_dict_1[key]:
                    match_level_found = True
                    break
            if not match_level_found:
                result_dict_1[key] = "No match level found"

        return result_dict_1,result_dict_2,result_dict_3        
        
    key_column = 'Peak_name_1'
    value_columns = ['pD1RD sum_1', 'pD1RD sum_2']
    alt_value_column = 'Peak_name_2'
    diff_column = 'Absolute Difference of Total pD1RD'
    check_column = 'Match Type_j'




    result_dict_type_j = create_dict_match_type_j(deal_mummer_data, key_column, alt_value_column)
    deal_mummer_data['Match Type_j'] = deal_mummer_data.apply(fill_absolute_difference, axis=1, args=(key_column, result_dict_type_j))

    result_dict = create_dict_diff(deal_mummer_data, key_column, value_columns, check_column,alt_value_column,)
    result_dict_type = create_dict_match_type(deal_mummer_data, key_column, alt_value_column)


    deal_mummer_data['Absolute Difference of Total pD1RD'] = deal_mummer_data.apply(fill_absolute_difference, axis=1, args=(key_column, result_dict))
    deal_mummer_data['Match Type'] = deal_mummer_data.apply(fill_absolute_difference, axis=1, args=(key_column, result_dict_type))

    result_dict_level = create_dict_match_level(deal_mummer_data, key_column, diff_column)
    #print(result_dict_level)
    deal_mummer_data['Match level'] = deal_mummer_data.apply(fill_absolute_difference, axis=1, args=(key_column, result_dict_level))
    
    # result_dict_level_1,result_dict_level_2,result_dict_level_3 = create_dict_match_level_many(deal_mummer_data, key_column, diff_column)
    #print(result_dict_level)
    deal_mummer_data['Match level'] = deal_mummer_data.apply(fill_absolute_difference, axis=1, args=(key_column, result_dict_level))


    dict_no_peak = create_dict_no_peak(deal_mummer_data,key_column,'Match Type')

    #print(df)
    
    
    # assuming 'df' is your DataFrame and 'Match Type_j' is a column in that DataFrame
    def process_row(row):
        match_type_j = row['Match Type_j']
        
        if pd.isna(match_type_j) or match_type_j == '':
            return np.nan
        elif match_type_j in ['GAP', 'No peak']:
            return '-'  # return '-' if the data is 'GAP' or 'No peak'
        else:
            parts = match_type_j.split('-')
            #print(parts)
            if 'H' in parts[0] and parts[0] == parts[1]:
                return 'same bin'  # return 'the same bin' if all parts are the same and contain 'H'
            else:
                return 'different bin'  # return 'different bin' otherwise

    deal_mummer_data['bin type'] = deal_mummer_data.apply(process_row, axis=1)

    
    def compare_dicts(dict1, dict2):
        different_values = {}
        for key in dict1:
            if key in dict2:
                if dict1[key] != dict2[key]:
                    different_values[key] = (dict1[key], dict2[key])
        return different_values

    def create_dict_from_dataframe_peak_2(df):
        result_dict = {}
        for _, row in df.iterrows():
            key = row['Peak_name_2']
            value = row['Peak_name_1']

            if isinstance(key, str):
                key = key.strip()
            if isinstance(value, str):
                value = value.strip()

            if (not key or key == "") or (not value or value == ""):
                continue

            if key not in result_dict:
                result_dict[key] = [value]
            else:
                if value not in result_dict[key]:
                    result_dict[key].append(value)
        return result_dict



    dict_3 =create_dict_from_dataframe_peak_2(deal_mummer_data)

    def process_alternate_data(df, condition):
        filtered_df = df[df['Peak_name_1'] == condition]
        num_copy_row_2 = (filtered_df['flag_copy_2'] == 'copy_row_2').sum()
        if num_copy_row_2 != 1:
            copy_row_2_indexes = filtered_df[filtered_df['flag_copy_2'] == 'copy_row_2'].index
            copy_row_2_indexes = copy_row_2_indexes[:-1]
            df = df.drop(copy_row_2_indexes)
            df = df.reset_index(drop=True)        
        return df


    #print(dict_1)
    #print(dict_2)
    dict_1 = create_dict_from_dataframe(deal_mummer_data)
    for key, value in dict_1.items():
        if isinstance(value, list):
            all_elements_empty = all(
                (elem == '' or elem is None or (isinstance(elem, float) and math.isnan(elem))) for elem in value
            )
            if not value or all_elements_empty:  
                print("\n" + "---------------Unmatched peak---------------------------")
                print(f"Key: {key}, Value: {value}")
                print("--------------------------------------------------------" + "\n")
            if len(value) > 1:
                deal_mummer_data = process_alternate_data(deal_mummer_data, key)

    columns_to_modify = ['Representative main motif type_1', 'Representative main motif type_2']  
    import numbers

    def not_empty(value):
        return (
            pd.notna(value)
            and value != ""
            and not (isinstance(value, str) and value.isspace())
            and not (isinstance(value, numbers.Number) and value == 0)
        )
        
        
    import re

    def insert_n_after_bracket(s):
        match = re.search(r'\](\w)', s)
        if match:
            index = match.start(1)
            return s[:index] + 'n' + s[index:]
        else:
            return s + 'n'


    for column in columns_to_modify:
        deal_mummer_data[column] = deal_mummer_data[column].apply(
            lambda x: insert_n_after_bracket(str(x)) if not_empty(x) and (str(x).endswith(']') or str(x) != '[MOTIFmix]') else x
        )


    def calculate_interval_size(row):
        if row["Extra_Column"] == "Inserted":
            interval_size_s1 = int(row["[E1]"] - row["[S1]"] + 2)
            interval_size_s2 = int(row["[E2]"] - row["[S2]"] + 2)

            row["[S1]"] = f"Interval Size: {interval_size_s1}"
            row["[E1]"] = f"Interval Size: {interval_size_s1}"
            row["[LEN1]"] = f"Interval Size: {interval_size_s1}"

            row["[S2]"] = f"Interval Size: {interval_size_s2}"
            row["[E2]"] = f"Interval Size: {interval_size_s2}"
            row["[LEN2]"] = f"Interval Size: {interval_size_s2}"

        return row
        
    def filter_fn(match_type):
        if not match_type or pd.isna(match_type):
            return False

        parts = match_type.split("-")

        if len(parts) != 2:
            return False

        return parts[0] == parts[1]

        
    new_column_order  = [
        "Representative main motif type_1","ucsc_name_1", "Start_1", "End_1", "[S1]", "[E1]", "[LEN1]", "[%IDY]","[LEN2]", "[S2]", "[E2]", "Start_2","End_2",
        "ucsc_name_2","Representative main motif type_2","pD1RD sum_1", "pD1RD sum_2",'Absolute Difference of Total pD1RD','Match level','bin type',
        'Match Type','Match Type_j','Peak_name_1','Peak_name_2', "Motif type_1", "flag_copy_1","index_no", "Extra_Column", "Motif type_2","flag_copy_2",
    ]

    deal_mummer_data_new = deal_mummer_data[new_column_order]
    
    
    #----------------------------------------------Filter data for Hn-Hn-------------------------------------------------------------------

    filtered_deal_mummer_data = deal_mummer_data_new[deal_mummer_data_new["Match Type_j"].apply(filter_fn)]

    unique_count = filtered_deal_mummer_data["Peak_name_1"].nunique()


    print("Unique count of 'Peak_name_1'-"f'{chr_parm}'": "f'{unique_count}')
    
    df_refseq_mrnas =pd.read_excel("RefSeq_mRNAs-Mark\\"+"CHM13-HMDP_all-exon-intron-07-12.xlsx", engine='openpyxl')

    subset_df = filtered_deal_mummer_data.copy()

    subset_df.loc[:, 'RefSeq_mRNAs-Mark'] = None
    for idx, row in subset_df.iterrows():
        #print(row['Peak_name_1'])
        matching_rows = df_refseq_mrnas[df_refseq_mrnas['Peak_name'] == row['Peak_name_1']]
        #print(matching_rows)
        if not matching_rows.empty:
            subset_df.at[idx, 'RefSeq_mRNAs-Mark'] = matching_rows.iloc[0]['RefSeq mRNAs-Mark']
            
    # Custom aggregate function
    def custom_agg(group):
        row_with_value = group[group['Peak_name_2'].apply(lambda x: x not in [None, '', np.nan])]
        if not row_with_value.empty:
            return row_with_value.iloc[0]
        else:
            return group.iloc[-1]

    # Applying custom aggregation functions using groupby() and apply()
    subset_df_result = subset_df.groupby('Peak_name_1').apply(custom_agg).reset_index(drop=True)


    subset_df_result.to_excel(otherfile+"filtered_data_"+chr_parm+".xlsx", index=False, engine='openpyxl')
    #-----------------------------------------------Filter data for Hn-Hn------------------------------------------------------------------
    
    #----------------------------------------------Counting of different types-------------------------------------------------------------------    
    def classify_type(match_type):
        try:
            if pd.isna(match_type) or match_type == 'GAP':
                return 'GAP'
            if match_type == 'No peak':
                return 'No peak'
            parts = match_type.split("-")
            if parts[0] == parts[1]:
                return 'Hn-Hn'
            if 'H' in parts[1]:
                return 'Hn-Hm'
            if 'M' in parts[1] and 'H' not in parts[1]:
                return 'H-M'
            if 'L' in parts[1] and 'H' not in parts[1] and'M' not in parts[1]:
                return 'H-L'                
        except Exception as e:
            print(f"Error: {e}")
            print(f"Parts: {parts}")
            return None
    df_type_count_1 = deal_mummer_data_new.drop_duplicates(subset='Peak_name_1', keep='first')
    df_type_count = df_type_count_1.drop_duplicates(subset='Peak_name_1', keep='last')
    duplicated_rows = df_type_count[df_type_count.duplicated('Peak_name_1', keep=False)]
    if not duplicated_rows.empty:
        print("Duplicate rows found:")
        print(duplicated_rows)
    df_type_count =df_type_count.copy()
    df_type_count.loc[:, 'category'] = df_type_count['Match Type_j'].apply(classify_type)
    counts_type = df_type_count['category'].value_counts()
    if 'GAP' in counts_type:
        counts_type['GAP'] -= 1

    counts_df = pd.DataFrame(counts_type).transpose()
    sum_counts_type = counts_type.sum()
    last_peak_name = df_type_count['Peak_name_1'].iloc[-1]
    number_to_compare = last_peak_name.split('-')[2].split('HMDP')[1].split('m')[0]
    if sum_counts_type != int(number_to_compare):
        print(f"Difference found: sum of counts_type is {sum_counts_type}, but number in Peak_name_1 is {number_to_compare}")
    counts_df['peak_sum'] = sum_counts_type
    counts_df['peak_sum_ture'] = number_to_compare  # Add the file name to the dataframe
    counts_df['chr_no'] = chr_parm  # Add the file name to the dataframe
    result_df_count = pd.concat([result_df_count, counts_df])
    #----------------------------------------------Counting of different types-------------------------------------------------------------------        
    
    

    deal_mummer_data_new = deal_mummer_data_new.copy()
    deal_mummer_data_new.loc[:, 'RefSeq_mRNAs-Mark'] = ''
    for idx, row in deal_mummer_data_new.iterrows():
        #print(row['Peak_name_1'])
        matching_rows = df_refseq_mrnas[df_refseq_mrnas['Peak_name'] == row['Peak_name_1']]
        #print(matching_rows)
        if not matching_rows.empty:
            deal_mummer_data_new.at[idx, 'RefSeq_mRNAs-Mark'] = matching_rows.iloc[0]['RefSeq mRNAs-Mark']    
        

    deal_mummer_data_new = deal_mummer_data_new.apply(calculate_interval_size, axis=1)

    #df.to_csv("c13_g38_chr01_comparison_2_interval.csv", index=False)
    # Save DataFrame to an xlsx file
    with pd.ExcelWriter(infile+"c13_g38_chr"+chr_parm+"_comparison.xlsx", engine='openpyxl') as writer:
        deal_mummer_data_new.to_excel(writer, index=False, sheet_name="Sheet1")



    def merge_rows_and_columns(worksheet, column_indices, start_row, end_row):
        def is_merged_cell(cell):
            for merged_cells in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_cells:
                    return True
            return False

        
        merge_ranges = []
        for row in range(start_row, end_row + 1):
            for col in column_indices:
                cell = worksheet.cell(row=row, column=col)
                if is_merged_cell(cell):
                    continue

                current_value = cell.value
                row_end = row
                col_end = col

                
                for r in range(row + 1, end_row + 1):
                    next_cell = worksheet.cell(row=r, column=col)
                    if is_merged_cell(next_cell) or next_cell.value != current_value:
                        break
                    row_end = r

                
                for c in column_indices[column_indices.index(col) + 1:]:
                    next_cell = worksheet.cell(row=row, column=c)
                    if is_merged_cell(next_cell) or next_cell.value != current_value:
                        break
                    col_end = c

                if row_end != row or col_end != col:
                    merge_ranges.append((row, col, row_end, col_end))

        
        for range_info in merge_ranges:
            row, col, row_end, col_end = range_info
            worksheet.merge_cells(start_row=row, start_column=col, end_row=row_end, end_column=col_end)
            merged_cell = worksheet.cell(row=row, column=col)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    def can_merge_cells_without_empty(cell1, cell2, merged_cells_dict):
        if cell1.coordinate in merged_cells_dict or cell2.coordinate in merged_cells_dict:
            return False

        if cell1.value is None or cell1.value == '' or cell2.value is None or cell2.value == '':
            return False

        return cell1.value == cell2.value

    def merge_rows_and_columns_without_empty_v3(worksheet, column_indices, start_row, end_row):
        merged_cells_dict = {}

        for merged_cells in worksheet.merged_cells.ranges:
            for cell in merged_cells:
                merged_cells_dict[cell.coordinate] = True

        for col in column_indices:
            row_start = start_row

            while row_start <= end_row:
                cell1 = worksheet.cell(row=row_start, column=col)

                if cell1.coordinate in merged_cells_dict or cell1.value is None or cell1.value == '':
                    row_start += 1
                    continue

                row_end = row_start + 1
                while row_end <= end_row:
                    cell2 = worksheet.cell(row=row_end, column=col)

                    if cell2.value is None or cell2.value == '':
                        break

                    if not can_merge_cells_without_empty(cell1, cell2, merged_cells_dict):
                        break

                    row_end += 1

                if row_end > row_start + 1:
                    worksheet.merge_cells(start_row=row_start, start_column=col, end_row=row_end - 1, end_column=col)
                    merged_cell = worksheet.cell(row=row_start, column=col)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                    for row in range(row_start + 1, row_end):
                        merged_cells_dict[worksheet.cell(row=row, column=col).coordinate] = True

                row_start = row_end
                
                
    from openpyxl.utils import get_column_letter
    def merge_rows_based_on_another_column(worksheet, column_indices, reference_column_indices, start_row, end_row):
        merged_cells_ranges = []

        for ref_col in reference_column_indices:
            for merged_cells in worksheet.merged_cells.ranges:
                if get_column_letter(ref_col) in merged_cells.coord:
                    merged_cells_ranges.append(merged_cells)

        for col, ref_col in zip(column_indices, reference_column_indices):
            for merged_cells in merged_cells_ranges:
                min_row, min_col, max_row, max_col = merged_cells.min_row, merged_cells.min_col, merged_cells.max_row, merged_cells.max_col

                if min_col == ref_col:
                    worksheet.merge_cells(start_row=min_row, start_column=col, end_row=max_row, end_column=col)
                    merged_cell = worksheet.cell(row=min_row, column=col)
                    merged_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    wb = load_workbook(infile+"c13_g38_chr"+chr_parm+"_comparison.xlsx")
    ws = wb.active




    columns_to_merge_1 = ["[S1]", "[E1]", "[LEN1]", "[%IDY]", "[LEN2]","[S2]", "[E2]"]

    columns_to_merge_2 = [
        "Peak_name_1", "Start_1", "End_1", "Motif type_1",
        "Start_2", "End_2", "Motif type_2",

    ]


    column_name_to_index = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    column_indices = [column_name_to_index[col_name] for col_name in columns_to_merge_1]
    merge_rows_and_columns(ws, column_indices, 2, ws.max_row)



    rep_main_motif_type_cols = ["Representative main motif type_1", "Representative main motif type_2"]
    peak_name_cols = ["Peak_name_1", "Peak_name_2"]
    rep_main_motif_type_cols_p =["pD1RD sum_1","pD1RD sum_2"]
    peak_name_cols_1 = ["Peak_name_1","Peak_name_1","Peak_name_1","Peak_name_1","Peak_name_1","Peak_name_2","Peak_name_1","Peak_name_1"]
    rep_diff = ['Absolute Difference of Total pD1RD',"Match Type","Match level","Match Type_j","ucsc_name_1","ucsc_name_2","RefSeq_mRNAs-Mark",'bin type']
    
    column_indices_other = [column_name_to_index[col_name] for col_name in columns_to_merge_2]
    


    merge_rows_and_columns_without_empty_v3(ws, column_indices_other, 2, ws.max_row) 

    def merge_peak_name_2_based_on_conditions_and_adjacent_values(worksheet):
        merged_ranges_for_peak_name_1 = []

        # Identify rows that are merged based on Peak_name_1
        for merged_cells in worksheet.merged_cells.ranges:
            if get_column_letter(column_name_to_index["Peak_name_1"]) in merged_cells.coord:
                merged_ranges_for_peak_name_1.append((merged_cells.min_row, merged_cells.max_row))

        col = column_name_to_index["Peak_name_2"]
        

        values = [cell.value for cell in worksheet[f"{get_column_letter(col)}"]]

        row_start = 2  # Assuming header is at row 1

        while row_start <= worksheet.max_row:
            peak_name_2_value = values[row_start - 1]

            # If Peak_name_2 is "GAP" or "No peak"
            if peak_name_2_value in ["GAP", "No peak"]:
                corresponding_peak_1_range = None
                for min_row, max_row in merged_ranges_for_peak_name_1:
                    if min_row <= row_start <= max_row:
                        corresponding_peak_1_range = (min_row, max_row)
                        break
                
                # If found a corresponding merged range in Peak_name_1
                if corresponding_peak_1_range:
                    min_row, max_row = corresponding_peak_1_range
                    # If all values in the range are the same for Peak_name_2
                    if all(v == peak_name_2_value for v in values[min_row-1:max_row]):
                        worksheet.merge_cells(start_row=min_row, start_column=col, end_row=max_row, end_column=col)
                        row_start = max_row + 1
                    else:
                        row_start += 1
                else:
                    row_start += 1

            # If Peak_name_2 value is not empty, not "GAP", not "No peak" and adjacent rows are the same
            elif peak_name_2_value:
                row_end = row_start
                while row_end <= worksheet.max_row and values[row_end - 1] == peak_name_2_value:
                    row_end += 1

                if row_end - row_start > 1:
                    worksheet.merge_cells(start_row=row_start, start_column=col, end_row=row_end - 1, end_column=col)
                    row_start = row_end
                else:
                    row_start += 1
            else:
                row_start += 1

    # Calling the function
    merge_peak_name_2_based_on_conditions_and_adjacent_values(ws)




    column_indices_rep_main_motif_type = [column_name_to_index[col_name] for col_name in rep_main_motif_type_cols]
    column_indices_peak_name = [column_name_to_index[col_name] for col_name in peak_name_cols]
    column_indices_rep_pdrd = [column_name_to_index[col_name] for col_name in rep_main_motif_type_cols_p]
    
    column_indices_peak_name_1 = [column_name_to_index[col_name] for col_name in peak_name_cols_1]
    column_indices_rep_diff = [column_name_to_index[col_name] for col_name in rep_diff]
    merge_rows_based_on_another_column(ws, column_indices_rep_main_motif_type, column_indices_peak_name, 2, ws.max_row)
    
    merge_rows_based_on_another_column(ws, column_indices_rep_pdrd, column_indices_peak_name, 2, ws.max_row)
    
    merge_rows_based_on_another_column(ws, column_indices_rep_diff, column_indices_peak_name_1, 2, ws.max_row)
    wb.save(outfile+"c13_g38_chr"+chr_parm+"_comparison_complete.xlsx")


    def find_column_indices_by_headers(ws, headers):
        header_row = ws[1]
        column_indices = []
        for header in headers:
            for cell in header_row:
                if cell.value == header:
                    column_indices.append(cell.column - 1)
                    break
        return column_indices

    def replace_partial_text(ws, column_headers, texts_to_replace, replacement=""):
        column_indices = find_column_indices_by_headers(ws, column_headers)
        for row in ws.iter_rows(min_row=2):  
            for column_index in column_indices:
                cell = row[column_index]
                if cell.value is not None:
                    for text_to_replace in texts_to_replace:
                        if text_to_replace in str(cell.value):
                            cell.value = str(cell.value).replace(text_to_replace, replacement)


    
    wb.save(lastfile+"c13_g38_chr"+chr_parm+"_comparison_complete_last.xlsx")
    print("c13_g38_chr"+f'{chr_parm}'+"_comparison_complete_last.xlsx"+"------File generated")
    

filtered_df['RefSeq_mRNAs-Mark'] = None
for idx, row in filtered_df.iterrows():
    matching_rows = df_refseq_mrnas[df_refseq_mrnas['Peak_name'] == row['Peak_name_1']]
    if not matching_rows.empty:
        filtered_df.at[idx, 'RefSeq_mRNAs-Mark'] = matching_rows.iloc[0]['RefSeq mRNAs-Mark']
        
result_df_count.to_excel(lastfile+"result_df_count_all.xlsx", engine='openpyxl')
    
print("-------------------------------------------------------"+"\n")
print("Processing completed ----------------------Software version：V1.0")
input()